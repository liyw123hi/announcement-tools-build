#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
公告下载器 v2.2 - Windows 10 兼容版
功能：读取 Excel（产品名称 + 公告链接），逐一下载 zip/rar/7z 等压缩包，
      并以产品名称重命名，支持断点续传。
      下载完成后，可与指定目录内的压缩包进行内容一致性比对。

用法：
  python download_announcements_win10.py              # 仅下载
  python download_announcements_win10.py --compare    # 下载后执行比对
  python download_announcements_win10.py --verify     # 仅执行比对（跳过下载）
配置文件：config.txt（与本脚本同目录）

Windows 10 安装依赖：
  pip install openpyxl rarfile py7zr
"""

import os
import sys
import re
import json
import time
import hashlib
import configparser
import urllib.request
import urllib.error
import urllib.parse
import zipfile
import io
import tempfile
import shutil

# Windows 兼容性：确保使用正确的路径分隔符
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.txt")

# Windows 临时目录（替代 /tmp）
TEMP_DIR = os.environ.get('TEMP', os.path.join(SCRIPT_DIR, 'temp'))


def load_config():
    """读取 config.txt 配置"""
    cfg = configparser.ConfigParser()
    if not os.path.exists(CONFIG_FILE):
        print(f"错误：找不到配置文件 {CONFIG_FILE}")
        sys.exit(1)
    cfg.read(CONFIG_FILE, encoding="utf-8")

    def get(key, default=""):
        return cfg.get("DEFAULT", key).strip() if cfg.has_option("DEFAULT", key) else default

    def path(p):
        if os.path.isabs(p):
            return p
        return os.path.join(SCRIPT_DIR, p)

    # 读取公告类型关键字（用于比对时识别日期差异）
    anno_keywords_raw = get("anno_keywords", "")
    if anno_keywords_raw:
        anno_keywords = [k.strip() for k in anno_keywords_raw.split(",") if k.strip()]
    else:
        anno_keywords = [
            "净值公告", "临时公告", "兑付公告", "成立公告",
            "激励公告", "加米公告", "报酬公告", "说明书",
            "扭赏公告", "预售公告", "兑回公告", "分配公告",
            "免贷公告", "原则公告", "增配公告", "缴费公告",
            "公告書", "公告书", "招商手册", "预售说明",
            "扭赏说明", "临时说明", "最新公告", "兑付说明",
            "支付公告",
        ]

    config = {
        "excel_path":  path(get("excel_path", "季度报告.xlsx")),
        "save_dir":    path(get("save_dir", "./downloads")),
        "compare_dir": path(get("compare_dir", "")),
        "download_interval": float(get("download_interval") or 2.0),
        "proxy":       get("proxy"),
        "timeout":     int(get("timeout") or 30),
        "state_file":  path(get("state_file", ".download_state.json")),
        "checksum":    get("checksum", "sha256").lower(),  # sha256 / md5 / size
        # 运行模式：1=仅下载  2=仅比对  3=下载+比对
        "mode":        int(get("mode", "1")),
        # 比对时使用的公告类型关键字（用于识别日期差异）
        "anno_keywords": anno_keywords,
    }

    for k in ["download_interval", "timeout"]:
        try:
            config[k] = int(config[k])
        except ValueError:
            config[k] = 30 if k == "timeout" else 2.0

    return config


# ─────────────────────────────────────────────────────────────────────────────
#  文件名清洗
# ─────────────────────────────────────────────────────────────────────────────
def sanitize_filename(name, max_len=180):
    illegal_chars = r'[\\/:*?"<>|]'
    name = re.sub(illegal_chars, "_", name).strip()
    if len(name) > max_len:
        name = name[:max_len]
    name = re.sub(r'_+', '_', name).strip('_')
    return name or "未命名产品"


def fmt_size(num_bytes):
    """将字节数格式化为可读字符串，自动选择 B / KB / MB / GB"""
    if num_bytes is None or num_bytes < 0:
        return "未知"
    if num_bytes >= 1024 * 1024 * 1024:
        return f"{num_bytes / (1024**3):.2f} GB"
    if num_bytes >= 1024 * 1024:
        return f"{num_bytes / (1024**2):.2f} MB"
    if num_bytes >= 1024:
        return f"{num_bytes / 1024:.1f} KB"
    return f"{num_bytes} B"


# ─────────────────────────────────────────────────────────────────────────────
#  Excel 读取
# ─────────────────────────────────────────────────────────────────────────────
def read_excel(excel_path):
    if not os.path.exists(excel_path):
        print(f"错误：找不到 Excel 文件 {excel_path}")
        sys.exit(1)
    print(f"正在读取 Excel: {excel_path}")
    try:
        import openpyxl
    except ImportError:
        print("缺少 openpyxl 模块，请先安装：pip install openpyxl")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or all(v is None for v in row):
            continue
        name = str(row[0]).strip() if row[0] else ""
        url  = str(row[1]).strip() if row[1] else ""
        if name and url:
            records.append((name, url))
    wb.close()
    print(f"共读取 {len(records)} 条记录")
    return records


# ─────────────────────────────────────────────────────────────────────────────
#  下载器
# ─────────────────────────────────────────────────────────────────────────────
class Downloader:

    SUPPORTED_EXTS = {".zip", ".rar", ".7z", ".pdf",
                      ".tar", ".gz", ".bz2", ".tgz"}

    def __init__(self, config):
        self.cfg        = config
        self.save_dir   = config["save_dir"]
        self.interval   = config["download_interval"]
        self.timeout    = config["timeout"]
        self.state_file = config.get("state_file") or ""
        self.state      = self._load_state()

        os.makedirs(self.save_dir, exist_ok=True)

        # Windows 兼容的 User-Agent
        self.headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Referer": "https://www.ccb.com/",
            "Accept": (
                "application/zip, application/x-rar-compressed, "
                "application/x-7z-compressed, application/octet-stream, */*"
            ),
            "Accept-Encoding": "identity",
        }

        handlers = []
        if config.get("proxy"):
            handlers.append(
                urllib.request.ProxyHandler({
                    "http":  config["proxy"],
                    "https": config["proxy"],
                })
            )
        self.opener = urllib.request.build_opener(*handlers)
        self.opener.addheaders = list(self.headers.items())

    def _load_state(self):
        if not self.state_file or not os.path.exists(self.state_file):
            return {}
        try:
            with open(self.state_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}

    def _save_state(self):
        if not self.state_file:
            return
        try:
            with open(self.state_file, "w", encoding="utf-8") as f:
                json.dump(self.state, f, ensure_ascii=False, indent=2)
        except IOError as e:
            print(f"  [警告] 保存断点状态失败: {e}")

    def _get_ext(self, url):
        parsed = urllib.parse.urlparse(url)
        path = urllib.parse.unquote(parsed.path)
        ext = os.path.splitext(path)[1].lower()
        # 处理 .tar.gz 等双扩展名
        if ext == ".gz" and path.endswith(".tar.gz"):
            return ".tar.gz"
        if ext == ".bz2" and path.endswith(".tar.bz2"):
            return ".tar.bz2"
        return ext if ext in self.SUPPORTED_EXTS else ".zip"

    def _dest_path(self, name, url):
        ext = self._get_ext(url)
        return os.path.join(self.save_dir, sanitize_filename(name) + ext)

    def download_one(self, name, url):
        dest = self._dest_path(name, url)
        key  = f"{name}|{url}"
        saved_size = self.state.get(key, 0)

        if os.path.exists(dest) and saved_size == 0:
            print(f"  [跳过] 已存在（完整）: {os.path.basename(dest)}")
            return True, "already_complete", dest

        print(f"  下载: {os.path.basename(dest)}")
        req = urllib.request.Request(url)
        for k, v in self.headers.items():
            req.add_header(k, v)

        try:
            if saved_size > 0:
                req.add_header("Range", f"bytes={saved_size}-")
                print(f"    [断点续传] 已下载 {fmt_size(saved_size)}，继续...")

            with self.opener.open(req, timeout=self.timeout) as resp:
                code = resp.getcode()
                if code not in (200, 206):
                    msg = f"HTTP {code}"
                    print(f"    [失败] {msg}")
                    return False, msg, dest

                cr = resp.headers.get("Content-Range", "")
                real_total = None  # 文件真实总大小（从 Content-Range 解析）
                if cr:
                    m = re.search(r"/(\d+)", cr)
                    if m:
                        real_total = int(m.group(1))
                        if saved_size > 0 and real_total == saved_size:
                            self.state[key] = 0
                            self._save_state()
                            print(f"    [完成] {os.path.basename(dest)}（断点续传检测到完整）")
                            return True, "complete", dest

                # 断点续传时 Content-Length 仅为本段大小，不用于进度计算
                content_length = resp.headers.get("Content-Length")
                # 优先用 real_total（Content-Range 总大小），否则用 Content-Length
                progress_total = real_total or (int(content_length) if content_length else None)
                mode = "ab" if saved_size > 0 else "wb"
                with open(dest, mode) as f:
                    downloaded = saved_size
                    last_pct = -1
                    while True:
                        chunk = resp.read(8192)
                        if not chunk:
                            break
                        f.write(chunk)
                        downloaded += len(chunk)
                        # 进度百分比变化时再打印（避免刷屏）
                        if progress_total:
                            pct = downloaded / progress_total * 100
                            pct_int = int(pct)
                            if pct_int != last_pct:
                                last_pct = pct_int
                                print(f"\r    进度: {downloaded}/{progress_total} ({pct:.1f}%)", end="")

                print()  # 等待下载真正完成后再换行
                final_size = os.path.getsize(dest)
                # 用 real_total（更准确）或 Content-Length 做校验
                expected = real_total or (int(content_length) if content_length else 0)
                if expected and final_size != expected:
                    print(f"    [警告] 文件大小不匹配（期望 {fmt_size(expected)}, 实际 {fmt_size(final_size)}）")

                self.state[key] = 0
                self._save_state()
                print(f"    [完成] {os.path.basename(dest)} ({fmt_size(final_size)})")
                return True, "complete", dest

        except urllib.error.HTTPError as e:
            msg = f"HTTP 错误: {e.code} {e.reason}"
            print(f"    [失败] {msg}")
            return False, msg, dest
        except urllib.error.URLError as e:
            msg = f"网络错误: {e.reason}"
            print(f"    [失败] {msg}")
            return False, msg, dest
        except Exception as e:
            msg = f"异常: {e}"
            print(f"    [失败] {msg}")
            return False, msg, dest

    def run(self, records):
        total = len(records)
        success = fail = skip = 0
        failed_records = []

        for i, (name, url) in enumerate(records, 1):
            print(f"\n[{i}/{total}]")
            ok, reason, dest = self.download_one(name, url)
            if ok:
                success += 1
                if reason == "already_complete":
                    skip += 1
            else:
                fail += 1
                failed_records.append((name, url, reason))

            if i < total and self.interval > 0:
                time.sleep(self.interval)

        print("\n" + "=" * 60)
        print("下载完成")
        print(f"  成功: {success}  |  跳过: {skip}  |  失败: {fail}")
        print(f"  保存目录: {self.save_dir}")
        print("=" * 60)

        if failed_records:
            print("\n失败记录：")
            for name, url, reason in failed_records:
                print(f"  - {name}")
                print(f"    URL: {url}")
                print(f"    原因: {reason}")

        return failed_records


# ─────────────────────────────────────────────────────────────────────────────
#  压缩包内容提取
# ─────────────────────────────────────────────────────────────────────────────
def _try_unrar(archive_path):
    """尝试用系统 unrar 命令提取 RAR 内容，返回 {文件名: SHA256} 或 None"""
    import subprocess
    unrar = shutil.which("unrar") or shutil.which("rar")
    if not unrar:
        return None

    tmp = io.BytesIO()
    try:
        result = subprocess.run(
            [unrar, "p", "-p-", "-c-", "-cfg-", archive_path],
            capture_output=True, timeout=120
        )
        if result.returncode == 0:
            tmp.write(result.stdout)
            tmp.seek(0)
            return tmp
    except (subprocess.SubprocessError, OSError):
        pass
    return None


def _checksum(data):
    """计算 SHA256"""
    return hashlib.sha256(data).hexdigest()


def _normalize_zip_name(filename):
    """
    ZIP 文件名在 Windows 中文环境下通常以 GBK 编码存储在 CP437 中。
    这里统一做 CP437→GBK 回退，避免文件名乱码导致比对失败。
    """
    raw = filename.encode("cp437", errors="replace")
    try:
        return raw.decode("gbk")
    except UnicodeDecodeError:
        return raw.decode("gb18030", errors="replace")


# 全局默认公告类型关键字（可被 config 覆盖）
_DEFAULT_ANNO_KEYWORDS = [
    "净值公告", "临时公告", "兑付公告", "成立公告",
    "激励公告", "加米公告", "报酬公告", "说明书",
    "扭赏公告", "预售公告", "兑回公告", "分配公告",
    "免贷公告", "原则公告", "增配公告", "缴费公告",
    "公告書", "公告书", "招商手册", "预售说明",
    "扭赏说明", "临时说明", "最新公告", "兑付说明",
    "支付公告",
]


def _looks_like_date_variant(extra_filename, archive_name, anno_keywords=None):
    """
    判断 extra_filename 是否是 archive_name 的公告文件变体。
    处理顺序（严格按此顺序，不可调换）：
      1. 去掉路径，只保留文件名
      2. 去掉扩展名
      3. 去掉基准名（archive_name）
      4. 与关键字列表比对，包含任意关键字则返回 True
    """
    keywords = anno_keywords or _DEFAULT_ANNO_KEYWORDS
    # 文件扩展名
    EXT_RE  = re.compile(r"\.(docx?|pdf|xlsx?|rar|zip|7z|txt)$", re.IGNORECASE)

    # 1. 去掉路径，只保留文件名
    name = os.path.basename(extra_filename)
    # 2. 去掉扩展名
    name_no_ext = EXT_RE.sub("", name)
    # 3. 去掉基准名
    if not name_no_ext.startswith(archive_name):
        return False
    remaining = name_no_ext[len(archive_name):]
    # 4. 与关键字比对
    for kw in keywords:
        if kw in remaining:
            return True
    return False


def _list_zip_contents(archive_path, mode="sha256"):
    """读取 zip 内部文件列表及校验和"""
    result = {}
    try:
        with zipfile.ZipFile(archive_path, "r") as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                data = zf.read(info.filename)
                # 用 GBK 解码兼容 Windows 中文文件名
                name = _normalize_zip_name(info.filename)
                if mode == "sha256":
                    result[name] = _checksum(data)
                elif mode == "md5":
                    result[name] = hashlib.md5(data).hexdigest()
                else:
                    result[name] = len(data)
    except Exception:
        pass
    return result


def _list_rar_contents(archive_path, mode="sha256"):
    """读取 rar 内部文件列表及校验和（Windows 优先用 rarfile 库）"""
    import subprocess

    # 方式一：系统 unrar 命令（如果安装了 WinRAR）
    unrar = shutil.which("unrar")
    if unrar:
        result = {}
        try:
            with tempfile.TemporaryDirectory() as td:
                out = subprocess.run(
                    [unrar, "x", "-o+", "-c-", "-cfg-", archive_path, td + "\\"],
                    capture_output=True, timeout=120
                )
                if out.returncode == 0:
                    for root, _, files in os.walk(td):
                        for fname in files:
                            fpath = os.path.join(root, fname)
                            rel = os.path.relpath(fpath, td)
                            with open(fpath, "rb") as f:
                                data = f.read()
                            if mode == "sha256":
                                result[rel] = _checksum(data)
                            elif mode == "md5":
                                result[rel] = hashlib.md5(data).hexdigest()
                            else:
                                result[rel] = len(data)
                return result
        except Exception:
            pass

    # 方式二：Python rarfile 库（Windows 推荐）
    try:
        import rarfile
        # 尝试找 unrar 工具路径
        unrar_tool = shutil.which("unrar") or shutil.which("rar")
        if unrar_tool:
            rarfile.UNRAR_TOOL = unrar_tool
        result = {}
        with rarfile.RarFile(archive_path) as rf:
            for entry in rf.infolist():
                if entry.isdir():
                    continue
                data = rf.read(entry)
                name = entry.filename
                # 尝试多种编码解码
                for enc in ("utf-8", "gbk", "gb2312", "gb18030", "cp437"):
                    try:
                        name = name.encode(enc).decode("utf-8", errors="strict")
                        break
                    except (UnicodeDecodeError, UnicodeEncodeError):
                        pass
                if mode == "sha256":
                    result[name] = _checksum(data)
                elif mode == "md5":
                    result[name] = hashlib.md5(data).hexdigest()
                else:
                    result[name] = len(data)
        return result
    except ImportError:
        print("  [提示] 安装 rarfile 以支持 RAR: pip install rarfile")
    except Exception:
        pass

    return {}


def _list_7z_contents(archive_path, mode="sha256"):
    """读取 7z 内部文件列表及校验和"""
    import subprocess

    # Windows 常见 7-Zip 安装路径
    seven_zip_paths = [
        shutil.which("7z"),
        shutil.which("7za"),
        r"C:\Program Files\7-Zip\7z.exe",
        r"C:\Program Files (x86)\7-Zip\7z.exe",
    ]
    exe = None
    for path in seven_zip_paths:
        if path and os.path.exists(path):
            exe = path
            break

    if not exe:
        # 尝试 Python py7zr 库
        try:
            import py7zr
            result = {}
            extract_tmp = os.path.join(TEMP_DIR, '_7z_extract_tmp')
            os.makedirs(extract_tmp, exist_ok=True)
            with py7zr.SevenZipFile(archive_path, "r") as z:
                z.extractall(path=extract_tmp)
                for root, _, files in os.walk(extract_tmp):
                    for fname in files:
                        fpath = os.path.join(root, fname)
                        rel = os.path.relpath(fpath, extract_tmp)
                        with open(fpath, "rb") as f:
                            data = f.read()
                        if mode == "sha256":
                            result[rel] = _checksum(data)
                        elif mode == "md5":
                            result[rel] = hashlib.md5(data).hexdigest()
                        else:
                            result[rel] = len(data)
            shutil.rmtree(extract_tmp, ignore_errors=True)
            return result
        except ImportError:
            print("  [提示] 安装 py7zr 以支持 7z: pip install py7zr")
        except Exception:
            pass
        return {}

    result = {}
    try:
        with tempfile.TemporaryDirectory() as td:
            out = subprocess.run(
                [exe, "x", f"-o{td}", "-y", archive_path],
                capture_output=True, timeout=120
            )
            if out.returncode == 0:
                for root, _, files in os.walk(td):
                    for fname in files:
                        fpath = os.path.join(root, fname)
                        rel = os.path.relpath(fpath, td)
                        with open(fpath, "rb") as f:
                            data = f.read()
                        if mode == "sha256":
                            result[rel] = _checksum(data)
                        elif mode == "md5":
                            result[rel] = hashlib.md5(data).hexdigest()
                        else:
                            result[rel] = len(data)
    except Exception:
        pass

    return result


def list_archive_contents(archive_path, mode="sha256"):
    """
    返回 dict：{ 文件名（含路径）: 校验值 }
    支持 .zip / .rar / .7z / .tar.gz / .tar.bz2
    解析失败时返回空 dict。
    """
    ext = os.path.splitext(archive_path)[1].lower()

    if ext == ".zip":
        return _list_zip_contents(archive_path, mode)

    if ext == ".rar":
        return _list_rar_contents(archive_path, mode)

    if ext == ".7z":
        return _list_7z_contents(archive_path, mode)

    if ext in (".tar.gz", ".tgz"):
        try:
            import tarfile
            result = {}
            with tarfile.open(archive_path, "r:gz") as tf:
                for member in tf.getmembers():
                    if not member.isfile():
                        continue
                    f = tf.extractfile(member)
                    if f is None:
                        continue
                    data = f.read()
                    name = member.name
                    if mode == "sha256":
                        result[name] = _checksum(data)
                    elif mode == "md5":
                        result[name] = hashlib.md5(data).hexdigest()
                    else:
                        result[name] = len(data)
            return result
        except Exception:
            return {}

    if ext in (".tar.bz2",):
        try:
            import tarfile
            result = {}
            with tarfile.open(archive_path, "r:bz2") as tf:
                for member in tf.getmembers():
                    if not member.isfile():
                        continue
                    f = tf.extractfile(member)
                    if f is None:
                        continue
                    data = f.read()
                    name = member.name
                    if mode == "sha256":
                        result[name] = _checksum(data)
                    elif mode == "md5":
                        result[name] = hashlib.md5(data).hexdigest()
                    else:
                        result[name] = len(data)
            return result
        except Exception:
            return {}

    if ext == ".tar":
        try:
            import tarfile
            result = {}
            with tarfile.open(archive_path, "r") as tf:
                for member in tf.getmembers():
                    if not member.isfile():
                        continue
                    f = tf.extractfile(member)
                    if f is None:
                        continue
                    data = f.read()
                    name = member.name
                    if mode == "sha256":
                        result[name] = _checksum(data)
                    elif mode == "md5":
                        result[name] = hashlib.md5(data).hexdigest()
                    else:
                        result[name] = len(data)
            return result
        except Exception:
            return {}

    # 不支持的格式
    return {}


# ─────────────────────────────────────────────────────────────────────────────
#  比对器
# ─────────────────────────────────────────────────────────────────────────────
class Comparator:

    def __init__(self, config):
        self.cfg      = config
        self.save_dir = config["save_dir"]
        self.compare_dir = config.get("compare_dir", "")
        self.checksum_mode = config.get("checksum", "sha256").lower()
        self.anno_keywords = config.get("anno_keywords", _DEFAULT_ANNO_KEYWORDS)

    def _match_file(self, product_name, base_dir):
        """
        在 base_dir 中找到与产品名称最匹配的文件。
        优先精确匹配，其次模糊匹配（产品名称含在文件名中）。
        """
        safe = sanitize_filename(product_name)
        if not os.path.isdir(base_dir):
            return None

        # 精确匹配（不含扩展名）
        for fname in os.listdir(base_dir):
            fbase = os.path.splitext(fname)[0]
            if sanitize_filename(fbase) == safe:
                return os.path.join(base_dir, fname)

        # 模糊匹配：产品名称关键字在文件名中
        keywords = re.sub(r'[^\w\u4e00-\u9fff]', '', product_name)
        best, best_score = None, 0
        for fname in os.listdir(base_dir):
            fbase = os.path.splitext(fname)[0]
            sfbase = sanitize_filename(fbase)
            if keywords in sfbase or sfbase in keywords:
                score = min(len(keywords), len(sfbase))
                if score > best_score:
                    best_score = score
                    best = os.path.join(base_dir, fname)
        return best

    def compare_one(self, product_name, downloaded_path):
        """
        比对已下载文件与对照目录中对应文件的内容。
        返回 (status, message):
          一致:   ("ok",      "内容完全一致")
          不一致: ("diff",    "内容不一致：..."   )
          找不到: ("no_ref",  "对照目录中未找到对应文件")
          错误:   ("error",   "错误信息")
        """
        ref_path = self._match_file(product_name, self.compare_dir)
        if not ref_path or not os.path.exists(ref_path):
            return "no_ref", f"对照目录未找到: {product_name}"

        ref_ext  = os.path.splitext(ref_path)[1].lower()
        down_ext = os.path.splitext(downloaded_path)[1].lower()
        if ref_ext != down_ext:
            return "error", (
                f"扩展名不一致：下载={down_ext}，对照={ref_ext}"
            )

        dl_list = list_archive_contents(downloaded_path, self.checksum_mode)
        ref_list = list_archive_contents(ref_path, self.checksum_mode)

        if not dl_list:
            return "error", "无法解析下载文件内容"
        if not ref_list:
            return "error", "无法解析对照文件内容"

        # 文件名校对
        dl_names  = set(dl_list.keys())
        ref_names = set(ref_list.keys())

        if dl_names != ref_names:
            missing_in_dl  = ref_names - dl_names
            missing_in_ref = dl_names  - ref_names
            newly_added   = []
            newly_missing = []
            date_added    = []
            date_missing  = []
            downloaded_base = os.path.splitext(os.path.basename(downloaded_path))[0]
            if missing_in_ref:
                newly_added = [f for f in missing_in_ref if not _looks_like_date_variant(f, downloaded_base, self.anno_keywords)]
                date_added  = [f for f in missing_in_ref if _looks_like_date_variant(f, downloaded_base, self.anno_keywords)]
            if missing_in_dl:
                newly_missing = [f for f in missing_in_dl if not _looks_like_date_variant(f, downloaded_base, self.anno_keywords)]
                date_missing = [f for f in missing_in_dl if _looks_like_date_variant(f, downloaded_base, self.anno_keywords)]
            msg = []
            if newly_added:
                msg.append(f"下载包有新增文件: {', '.join(sorted(newly_added))}")
            if newly_missing:
                msg.append(f"下载包缺少文件: {', '.join(sorted(newly_missing))}")
            if not msg:
                all_date = sorted(date_added + date_missing)
                names = ', '.join(all_date)
                return "ok", f"结果预期一致，仅日期差异: {names}"
            date_note = []
            if date_added:
                date_note.append(f"新增日期差异: {', '.join(sorted(date_added))}")
            if date_missing:
                date_note.append(f"缺失日期差异: {', '.join(sorted(date_missing))}")
            if date_note:
                msg.append("（另有日期差异仅供参考: " + "; ".join(date_note) + "）")
            return "diff", "; ".join(msg)

        # 内容校验
        diff_files = []
        for name in sorted(dl_names):
            if dl_list[name] != ref_list[name]:
                diff_files.append(name)

        if diff_files:
            return "diff", f"文件内容不一致: {', '.join(diff_files)}"

        return "ok", "结果预期一致"

    def run(self, records):
        if not self.compare_dir:
            print("错误：config.txt 中未设置 compare_dir（对照目录路径）")
            print("请在 config.txt 中添加：compare_dir = C:\\你的\\对照\\目录\\路径")
            return

        if not os.path.isdir(self.compare_dir):
            print(f"错误：对照目录不存在 {self.compare_dir}")
            return

        print(f"对照目录: {self.compare_dir}")
        print(f"校验方式: {self.checksum_mode.upper()}")
        print("=" * 60)

        results = {"ok": [], "diff": [], "no_ref": [], "error": []}
        total = len(records)

        for i, (name, url) in enumerate(records, 1):
            downloaded = os.path.join(self.save_dir, sanitize_filename(name) + os.path.splitext(url.split("/")[-1])[1].lower())
            # 实际文件扩展名从下载目录中查找
            ext = None
            for candidate in os.listdir(self.save_dir):
                if sanitize_filename(name) in sanitize_filename(candidate):
                    ext = os.path.splitext(candidate)[1].lower()
                    downloaded = os.path.join(self.save_dir, candidate)
                    break
            if not ext:
                print(f"[{i}/{total}] ? {name} -> 下载目录中未找到对应文件")
                results["no_ref"].append((name, "下载目录未找到"))
                continue

            status, msg = self.compare_one(name, downloaded)
            results[status].append((name, msg))

            icon = {"ok": "[OK]", "diff": "[DIFF]", "no_ref": "[SKIP]", "error": "[ERR]"}.get(status, "?")
            short_msg = msg[:60] + ("..." if len(msg) > 60 else "")
            print(f"[{i}/{total}] {icon} {name[:40]}")
            print(f"        {short_msg}")

        # 汇总
        print("\n" + "=" * 60)
        print("比对结果汇总")
        print(f"  [OK] 一致:   {len(results['ok'])} 个")
        print(f"  [DIFF] 不一致: {len(results['diff'])} 个")
        print(f"  [SKIP] 跳过:   {len(results['no_ref'])} 个（对照目录未找到）")
        print(f"  [ERR] 错误:  {len(results['error'])} 个")
        print("=" * 60)

        if results["diff"]:
            print("\n[DIFF] 不一致详情：")
            for name, msg in results["diff"]:
                print(f"  - {name}")
                print(f"    {msg}")

        if results["error"]:
            print("\n[ERR] 错误详情：")
            for name, msg in results["error"]:
                print(f"  - {name}: {msg}")

        # 写入 Excel 结果（追加 E、F 两列）
        excel_path = self.cfg.get("excel_path", "")
        if excel_path and os.path.exists(excel_path):
            try:
                from openpyxl.styles import Font, Border, Side, Alignment
                import openpyxl
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active

                def _thin_border():
                    s = Side(border_style="thin", color="000000")
                    return Border(left=s, right=s, top=s, bottom=s)

                # 找表头
                header = [c.value for c in ws[1]]
                col_result = None
                col_detail = None
                for i, h in enumerate(header):
                    if h == "对比结果":
                        col_result = i + 1
                    elif h == "详情":
                        col_detail = i + 1
                if col_result is None:
                    col_result = len(header) + 1
                    col_detail = len(header) + 2
                    ws.cell(1, col_result, "对比结果")
                    ws.cell(1, col_detail, "详情")

                # 表头加粗 + 加框
                for col in (col_result, col_detail):
                    cell = ws.cell(1, col)
                    cell.font = Font(bold=True)
                    cell.border = _thin_border()

                # 设置列宽
                ws.column_dimensions[ws.cell(1, col_result).column_letter].width = 18
                ws.column_dimensions[ws.cell(1, col_detail).column_letter].width = 35

                # 写入数据
                for i, (name, url) in enumerate(records, 2):
                    status, msg = None, ""
                    for key in ("ok", "diff", "no_ref", "error"):
                        for n, m in results.get(key, []):
                            if n == name:
                                status, msg = key, m
                                break
                        if status:
                            break
                    label = {"ok": "[OK] 结果预期一致", "diff": "[DIFF] 不一致",
                             "no_ref": "[SKIP] 跳过", "error": "[ERR] 错误"}.get(status, "?")
                    # 详情：去掉状态标签，只保留客观描述
                    detail = msg
                    if status == "ok":
                        if "仅日期差异" in msg:
                            import re as _re
                            _m = _re.search(r"仅日期差异:\s*(.+)$", msg)
                            if _m:
                                _files_str = _m.group(1)
                                _files = [f.strip() for f in _files_str.split(",")]
                                _remaining_parts = []
                                for _f in _files:
                                    _fname = os.path.basename(_f)
                                    _fname_no_ext = _re.sub(r'\.(docx?|pdf|xlsx?|rar|zip|7z|txt)$', '', _fname, flags=_re.I)
                                    if _fname_no_ext.startswith(name):
                                        _remaining = _fname_no_ext[len(name):]
                                    else:
                                        _remaining = _fname_no_ext
                                    for _kw in self.anno_keywords:
                                        if _kw in _remaining:
                                            if _remaining not in _remaining_parts:
                                                _remaining_parts.append(_remaining)
                                            break
                                if _remaining_parts:
                                    detail = "仅多了产品的" + "、".join(_remaining_parts)
                                else:
                                    detail = "仅多了日期相关的公告文件"
                            else:
                                detail = "仅多了日期相关的公告文件"
                        else:
                            detail = "内容完全一致"
                    elif status == "diff":
                        detail = msg.replace("内容不一致: ", "")
                    cell_r = ws.cell(i, col_result)
                    cell_d = ws.cell(i, col_detail)
                    cell_r.value = label
                    cell_d.value = detail
                    cell_r.border = _thin_border()
                    cell_d.border = _thin_border()
                wb.save(excel_path)
                print(f"\n结果已追加至 Excel: {excel_path}")
            except Exception as e:
                print(f"\nExcel 写入失败: {e}")

        return results


# ─────────────────────────────────────────────────────────────────────────────
#  主入口
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  公告下载器 v2.2 - Windows 10 兼容版")
    print("  支持 zip / rar / 7z / tar.gz 等压缩包")
    print("=" * 60 + "\n")

    config = load_config()
    mode = config.get("mode", 1)

    mode_names = {
        1: "仅下载",
        2: "仅比对",
        3: "下载 + 比对",
    }
    print(f"运行模式：{mode}（{mode_names.get(mode, '未知')}）")
    print("=" * 60 + "\n")

    # 读取 Excel（所有模式都需要）
    records = read_excel(config["excel_path"])
    if not records:
        print("Excel 中没有找到有效数据")
        sys.exit(0)

    failed = []

    # ── 模式 1 仅下载 ──
    if mode == 1:
        downloader = Downloader(config)
        failed = downloader.run(records)
        if failed:
            print("\n提示：重新运行本程序将自动从失败处继续下载（断点续传）")

    # ── 模式 2 仅比对 ──
    elif mode == 2:
        if not os.path.isdir(config.get("compare_dir", "")):
            print(f"错误：对照目录不存在或未设置 -> {config.get('compare_dir', '')}")
            sys.exit(1)
        Comparator(config).run(records)

    # ── 模式 3 下载 + 比对 ──
    elif mode == 3:
        if not os.path.isdir(config.get("compare_dir", "")):
            print(f"错误：对照目录不存在或未设置 -> {config.get('compare_dir', '')}")
            print("请在 config.txt 中正确填写 compare_dir")
            sys.exit(1)

        downloader = Downloader(config)
        failed = downloader.run(records)

        if failed:
            print("\n存在下载失败的文件，跳过比对。")
            print("修复后重新运行即可继续下载并比对。")
        else:
            print()
            Comparator(config).run(records)
        if failed:
            print("\n提示：重新运行本程序将自动从失败处继续下载（断点续传）")

    else:
        print(f"错误：mode={mode} 无效，仅支持 1 / 2 / 3")
        sys.exit(1)


if __name__ == "__main__":
    main()
